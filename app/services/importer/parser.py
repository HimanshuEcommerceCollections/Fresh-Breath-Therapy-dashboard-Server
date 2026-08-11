"""Reading an uploaded spreadsheet into headers + rows.

Handles the shapes real spreadsheets actually have rather than the tidy ones
documentation assumes: a title line above the headers, blank spacer rows,
duplicate column names, trailing empty columns from a stray formatted cell.

Everything is read into memory and nothing is written to blob storage. The
file contains PHI, and Cloudinary — where this app's other uploads go — has
no BAA covering it. The rows land in `import_rows` in Postgres, which is
already the system holding this data.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from openpyxl import load_workbook

# A cap, not a target. FBT's history is a few thousand rows; anything past
# this is a runaway file and should fail loudly rather than exhaust memory on
# a serverless function.
MAX_ROWS = 20_000
# How far down to look for the header line before giving up.
HEADER_SEARCH_DEPTH = 10
# Rows shown to the admin, and the sample the matcher reasons over.
SAMPLE_SIZE = 15


class SheetParseError(ValueError):
    """The file could not be read at all. Message is admin-facing."""


@dataclass
class ParsedSheet:
    headers: list[str]
    # {header: cell}, one per data row, in file order.
    rows: list[dict]
    # 1-based line numbers in the original file, parallel to `rows`, so an
    # error can name the line the admin will actually see when she opens it.
    row_numbers: list[int]
    header_row_number: int
    sheet_name: str | None = None

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    def samples(self, header: str, limit: int = SAMPLE_SIZE) -> list:
        """First non-empty values in one column."""
        out = []
        for row in self.rows:
            value = row.get(header)
            if value is not None and str(value).strip():
                out.append(value)
                if len(out) >= limit:
                    break
        return out

    def distinct(self, header: str, limit: int = 200) -> list[str]:
        """Distinct values in one column, most frequent first.

        Feeds the enum value-mapping screen. Also the only column content the
        LLM is ever shown for a low-cardinality field: a list of the six
        spellings of "ongoing" in a status column is a vocabulary, not a
        patient record.
        """
        counts: dict[str, int] = {}
        for row in self.rows:
            value = row.get(header)
            if value is None or not str(value).strip():
                continue
            key = str(value).strip()
            counts[key] = counts.get(key, 0) + 1
        ordered = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
        return [value for value, _ in ordered[:limit]]

    def value_counts(self, header: str) -> dict[str, int]:
        counts: dict[str, int] = {}
        for row in self.rows:
            value = row.get(header)
            if value is None or not str(value).strip():
                continue
            key = str(value).strip()
            counts[key] = counts.get(key, 0) + 1
        return counts


def _clean_header(value, index: int) -> str:
    if value is None or not str(value).strip():
        return f"Column {index + 1}"
    return " ".join(str(value).split())


def _dedupe(headers: list[str]) -> list[str]:
    """Duplicate column names are common and must stay distinguishable —
    they're the dict keys every later stage looks rows up by."""
    seen: dict[str, int] = {}
    out = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            out.append(f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            out.append(h)
    return out


def _score_header_row(cells: list) -> int:
    """How much a row looks like a header: short, non-numeric, mostly filled."""
    score = 0
    for cell in cells:
        if cell is None or not str(cell).strip():
            continue
        text = str(cell).strip()
        if len(text) > 60:
            continue          # a sentence — this is data, not a header
        try:
            float(text)
            continue          # a number — headers rarely are
        except ValueError:
            pass
        score += 1
    return score


def _find_header_row(grid: list[list]) -> int:
    """Index into `grid` of the most header-like row near the top.

    Sheets routinely open with a title ("FBT Client List 2024") and a blank
    line before the real headers; taking row 0 on faith would make the whole
    import look like one unmappable column.
    """
    best_index, best_score = 0, -1
    for i, row in enumerate(grid[:HEADER_SEARCH_DEPTH]):
        score = _score_header_row(row)
        if score >= 2 and score > best_score:
            best_index, best_score = i, score
    return best_index


def _build(grid: list[list], sheet_name: str | None = None) -> ParsedSheet:
    if not grid:
        raise SheetParseError("The file is empty.")

    header_index = _find_header_row(grid)
    raw_headers = grid[header_index]

    # Drop trailing empty columns — a single stray formatted cell far to the
    # right otherwise produces a dozen phantom "Column 14" entries.
    while raw_headers and (raw_headers[-1] is None or not str(raw_headers[-1]).strip()):
        raw_headers = raw_headers[:-1]
    if not raw_headers:
        raise SheetParseError("Couldn't find a header row with column names in it.")

    headers = _dedupe([_clean_header(h, i) for i, h in enumerate(raw_headers)])
    width = len(headers)

    rows: list[dict] = []
    row_numbers: list[int] = []
    for offset, cells in enumerate(grid[header_index + 1:], start=1):
        if all(c is None or not str(c).strip() for c in cells):
            continue  # blank spacer row
        padded = list(cells[:width]) + [None] * max(0, width - len(cells))
        rows.append(dict(zip(headers, padded)))
        # +1 twice: grid is 0-based, and spreadsheets number from 1.
        row_numbers.append(header_index + offset + 1)
        if len(rows) > MAX_ROWS:
            raise SheetParseError(
                f"The file has more than {MAX_ROWS:,} rows. Split it and import "
                "in parts."
            )

    if not rows:
        raise SheetParseError("Found column headers but no data rows beneath them.")

    return ParsedSheet(
        headers=headers,
        rows=rows,
        row_numbers=row_numbers,
        header_row_number=header_index + 1,
        sheet_name=sheet_name,
    )


def _parse_csv(content: bytes) -> ParsedSheet:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SheetParseError("Couldn't read the file's text encoding.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # a single-column file sniffs as nothing
    grid = [list(row) for row in csv.reader(io.StringIO(text), dialect)]
    return _build(grid)


def _parse_xlsx(content: bytes) -> ParsedSheet:
    try:
        # read_only streams rather than building the whole object graph;
        # data_only takes formula RESULTS, since a client-total column is
        # usually "=SUM(...)" and the formula text is useless to us.
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as exc:
        raise SheetParseError(f"Couldn't open the workbook: {exc}")

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return _build(grid, sheet_name=worksheet.title)
    finally:
        workbook.close()


def parse_sheet(content: bytes, filename: str) -> ParsedSheet:
    """Excel (.xlsx) and CSV only.

    Deliberately narrow, and enforced here rather than only in the browser's
    file picker — the picker's `accept` attribute is a filter, not a
    guarantee, and a drag-and-drop or a direct API call bypasses it entirely.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(content)
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    if name.endswith((".xls", ".xlsm")):
        raise SheetParseError(
            "Only .xlsx and .csv files are supported. Open this in Excel and "
            "use File → Save As → Excel Workbook (.xlsx), then upload that."
        )
    raise SheetParseError(
        f"Unsupported file type {filename!r}. Upload a .xlsx or .csv file."
    )
